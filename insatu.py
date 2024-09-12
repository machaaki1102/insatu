#import win32com.client as win32
#import pythoncom
import streamlit as st
import os


# # PDFに変換する関数
# def save_excel_as_pdf(excel_path, pdf_path, row_start, row_end, col_start, col_end):
#     pythoncom.CoInitialize()  # COMライブラリの初期化
#     excel = win32.Dispatch("Excel.Application")
#     excel.Visible = False
#     wb = excel.Workbooks.Open(excel_path)
#     ws = wb.Worksheets[0]
    
#     # 行と列を数値で動的に印刷範囲を設定
#     start_cell = ws.Cells(row_start, col_start).Address  # 開始セル
#     end_cell = ws.Cells(row_end, col_end).Address  # 終了セル
#     ws.PageSetup.PrintArea = f"{start_cell}:{end_cell}"  # 印刷範囲を設定
    
#     # PDFとして保存
#     ws.ExportAsFixedFormat(0, pdf_path)
#     wb.Close(SaveChanges=False)
#     excel.Quit()


# # 指定されたパスを確認
# file_path = "viv.xlsx"
# file_path_2 = "output.pdf"

# # ファイルの存在確認
# if os.path.exists(file_path):
#     # 絶対パスを取得して表示
#     absolute_path = os.path.abspath(file_path)
#     st.write(f"File exists at: {absolute_path}")
# else:
#     st.write("File does not exist at the specified path.")



# # StreamlitアプリのUIを作成
# st.title("Excel to PDF Converter")

# # Excelファイルのパスを入力またはファイル選択
# #excel_path = st.text_input("Enter Excel file path:", "viv.xlsx")
# #pdf_path = st.text_input("Enter output PDF file path:", "output.pdf")

# # 行と列の範囲を入力
# row_start = st.number_input("Start Row", min_value=1, value=1)
# row_end = st.number_input("End Row", min_value=1, value=25)
# col_start = st.number_input("Start Column (A=1, B=2, ...)", min_value=1, value=1)
# col_end = st.number_input("End Column (A=1, B=2, ...)", min_value=1, value=10)

# excel_path = os.path.abspath(file_path)
# pdf_path = os.path.abspath(file_path_2)


# # Excelファイルのパスが有効かどうかチェック
# if not os.path.exists(excel_path):
#     st.error("The specified Excel file does not exist. Please provide a valid path.")
# else:
#     # "Convert to PDF" ボタンをクリックしたらPDF変換処理を実行
#     if st.button("Convert to PDF"):
#         save_excel_as_pdf(excel_path, pdf_path, row_start, row_end, col_start, col_end)
#         st.success(f"Excel file has been converted to PDF and saved to {pdf_path}")

#         # 変換したPDFファイルのダウンロードリンクを表示
#         with open(pdf_path, "rb") as f:
#             pdf_data = f.read()
#         st.download_button(label="Download PDF", data=pdf_data, file_name=os.path.basename(pdf_path), mime='application/pdf')

import openpyxl

wb = openpyxl.load_workbook('bb_tem_finish_insatu.xlsx')
ws = wb['BB_テンプレ']

# 行数と列数を変数として定義
row_start = 1
row_end = 42
col_start = 1  # A列は1
col_end = 163    # E列は5

# 列番号をアルファベットに変換する
from openpyxl.utils import get_column_letter

col_start_letter = get_column_letter(col_start)
col_end_letter = get_column_letter(col_end)

# print_areaを設定
ws.print_area = f'{col_start_letter}{row_start}:{col_end_letter}{row_end}'

wb.save('output.xlsx')


with open('output.xlsx', 'rb') as file:
        mokuji_ekihi = file.read()

st.download_button(
        label="Download Excel File＜BB＞",  # ボタンのラベル
        data=mokuji_ekihi,  # ダウンロードするデータ
        file_name='output.xlsx',  # ダウンロード時のファイル名
        mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'  # MIMEタイプを指定
    )

ws.print_area = 'A1:E5'

wb.save('output.xlsx')